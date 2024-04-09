import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime  

pd.options.mode.copy_on_write = True

def process_begun():

    timestamp = datetime.now()
    str_date_time = timestamp.strftime("%Y-%m-%d @ %H:%M:%S")
    print("Invoice DM Source compilation began on", str_date_time)

def process_ended():

    timestamp = datetime.now()
    str_date_time = timestamp.strftime("%Y-%m-%d @ %H:%M:%S")
    print("Invoice DM Source compilation completed on", str_date_time)
    print(" ")

def eval_timestamp():
    timestamp = datetime.now()
    step_time = timestamp.strftime("%Y-%m-%d @ %H:%M:%S")
    print("step completed", step_time)

def compile_excel_files(folder_path, output_file):


    process_begun()

    

    # # Get a list of all Excel files in the folder
    # excel_files = [file for file in os.listdir(folder_path) if file.endswith(('.xls', '.xlsx'))]
    # print("excel file list retrieved")
    # eval_timestamp()

    # Check if there are any Excel files in the folder
    #if not excel_files:
     #   print("No Excel files found in the specified folder.")
      #  return

    #------------------------------------------------

    # Get a list of all Excel files in the directory
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

    # Sort the Excel files by size (largest first)
    excel_files.sort(key=lambda x: os.path.getsize(os.path.join(folder_path, x)), reverse=True)

    print("excel files sorted largest to smallest")
    eval_timestamp()

    # Initialize the base dataframe
    compiled_data = pd.DataFrame()
    print("compiled_data df created")
    eval_timestamp()

    # Read the first Excel file into the base dataframe
    first_file = True

    # Iterate over each Excel file in the directory
    for filename in excel_files:
        filepath = os.path.join(folder_path, filename)

        # Load the Excel workbook
        workbook = load_workbook(filepath)
        print(f"{filename} workbook loaded")
        eval_timestamp()
    
        # Get the active sheet
        sheet = workbook.active
    
        # Read the Excel file in chunks
        for row in sheet.iter_rows(values_only=True, max_row=sheet.max_row, max_col=sheet.max_column):
            chunk = pd.DataFrame([row], columns=[cell.value for cell in sheet[1]])
        
            # Remove rows that are duplicates with compiled_data
            if not first_file:
                chunk = chunk[~chunk.isin(compiled_data)].dropna()

            # Append the unique rows to the base dataframe
            compiled_data = pd.concat([compiled_data, chunk], ignore_index=True)

        if first_file:
            first_file = False

        # # Read the Excel file into a temporary dataframe
        # temp_df = pd.read_excel(filepath)
        # print("df completed")
        # eval_timestamp()
        
        # # Remove rows from temp_df that are duplicates with compiled_data
        # if not first_file:
        #     temp_df = temp_df[~temp_df.isin(compiled_data)].dropna()
        #     print(f"{filename} temp df evaluated")
        #     eval_timestamp()
        
        # # Append the temporary dataframe to the base dataframe
        # compiled_data = pd.concat([compiled_data, temp_df], ignore_index=True)
        # print("df appended")
        # eval_timestamp()

        # if first_file:
        #     first_file = False

    # Drop duplicate rows from the base dataframe
    #base_df.drop_duplicates(inplace=True)

    # Write the base dataframe to a CSV file
    #base_df.to_csv('output.csv', index=False)




    #-------------------------------------------------

    # # Initialize an empty list to store DataFrames
    # dfs = []
    # print("blank df created")
    # eval_timestamp()

    # # Loop through each Excel file and append its data to the list
    # for excel_file in excel_files:
    #     file_path = os.path.join(folder_path, excel_file)
    #     df = pd.read_excel(file_path) #, dtype=Inv_dtype_dict)
    #     dfs.append(df)
    #     print("df created")
    #     eval_timestamp()
    # # Concatenate the list of DataFrames into one
    # compiled_data = pd.concat(dfs, ignore_index=True)
    # print("data compiled")
    # eval_timestamp()

    # Convert date columns to datetime dtype

    

    date_columns = ['Last Updated Date', 'Payment Date', 'Date Received']

    compiled_data[date_columns] = compiled_data[date_columns].apply(lambda x: pd.to_datetime(x, format='%m/%d/%y'))
    print("dates converted")
    eval_timestamp()

    # Function to find the most recent date among three columns
    def find_latest_date(row):
        #return max(row['Created Date'], row['Last Updated Date'], row['Payment Date'], row['Date Received'])
        return max(row['Last Updated Date'], row['Payment Date'], row['Date Received'])
        
    # Apply the function row-wise to find the latest date
    compiled_data['Latest Record Date'] = compiled_data.apply(find_latest_date, axis=1)
    print("most recent records identified")
    eval_timestamp()
    #clean_data['Latest Record Date'] = clean_data.loc[:, date_columns].apply(find_latest_date, axis=1)

    #order by latest record date this will ensure latest records show first when manually viewing final source
    compiled_data = compiled_data.sort_values(by= 'Latest Record Date', ascending=False)
    print("data sorted")
    eval_timestamp()

    # Drop duplicates based on all columns
    compiled_data = compiled_data.drop_duplicates(subset=['Invoice ID', 'Latest Record Date'])
    print("duplicates dropped")
    eval_timestamp()

    # Write the compiled data to a new CSV file
    compiled_data.to_csv(output_file, index=False)
    print(f"Compiled data saved to {output_file}")
    eval_timestamp()

# Example usage
#folder_path = 'C:/Users/MatthewHieger/Documents/My Tableau Repository/Datasources/Coupa Datamart/Invoice Datamart files/'
#output_file = 'C:/Users/MatthewHieger/Documents/My Tableau Repository/Datasources/Coupa Datamart/Invoice DM Source.csv'

# Example usage in TEST folders
folder_path = 'C:/Users/MatthewHieger/Documents/My Tableau Repository/Datasources/Coupa Datamart/DATAMART COMPILE TEST/'
output_file = 'C:/Users/MatthewHieger/Documents/My Tableau Repository/Datasources/Coupa Datamart/DATAMART COMPILE TEST/Final Result/Invoice DM TEST Source.csv'

# Inv_dtype_dict = {'Invoice #': 'str',
# 'Supplier': 'str',
# 'Net Due Date': 'str',
# 'Total': 'float',
# 'Status': 'str',
# 'Delivery Method': 'str',
# 'Dispute Reason': 'str',
# 'Invoice Date': 'str',
# 'Invoice ID': 'str',
# 'Last Updated Date': 'str',
# 'Last Updated By': 'str',
# 'Payment Date': 'str',
# 'Payment Term': 'str',
# 'PO Number': 'str',
# 'PO ID': 'str',
# 'Requester': 'str',
# 'Supplier #': 'str',
# 'Chart Of Accounts': 'str',
# 'Created By': 'str',
# 'Created Date': 'str',
# 'Credit Applied': 'str',
# 'Current Approver': 'str',
# 'Date Received': 'str',
# 'Exported?': 'str',
# 'Last Exported At': 'str',
# 'Line Tags': 'str',
# 'Line Tolerance Failures': 'str',
# 'Original Invoice Number': 'str',
# 'Paid': 'str',
# 'Payment Channel': 'str',
# 'Payment Notes': 'str',
# 'Payment Num''s': 'str',
# 'Remit To Code': 'str',
# 'Resolved Line Tolerance Failures': 'str',
# 'Resolved Tolerance Failures': 'str',
# 'Source': 'str',
# 'Supplier Default Commodity': 'str',
# 'Supplier Total': 'float',
# 'Tags': 'str',
# 'Time with Current Approver': 'str',
# 'Tolerance Failures': 'str',
# 'Unanswered Comments': 'str',
# 'Latest Record Date': 'str'}

compile_excel_files(folder_path, output_file)

process_ended()